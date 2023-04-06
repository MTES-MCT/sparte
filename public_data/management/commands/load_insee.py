import logging

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from public_data.models import Commune, CommunePop
from public_data.models.mixins import TruncateTableMixin
from public_data.storages import DataStorage

logger = logging.getLogger("management.commands")


class TruncateComPop(TruncateTableMixin, CommunePop):
    class Meta:
        proxy = True


class Command(BaseCommand):
    help = "Charge les données de l'INSEE dans la BDD"

    def handle(self, *args, **options):
        logger.info("Start loading INSEE data")
        self.upload_pop()
        logger.info("End loading INSEE data")

    def get(self, remote_file_path, headers):
        logger.info(f"file={remote_file_path}")
        with DataStorage().open(remote_file_path) as file_stream:
            wb = load_workbook(file_stream, data_only=True)
        ws = wb.active
        return {r[0]: dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, max_col=len(headers), values_only=True)}

    def get_pop_data(self):
        remote_file_path = "base-pop-historiques-1876-2019.xlsx"
        headers = ["CODGEO", "LIBGEO"] + [f"P{year}" for year in range(2019, 2005, -1)]
        return self.get(remote_file_path, headers)

    def get_household_data(self):
        remote_file_path = "base-cc-coupl-fam-men-2018-lite.xlsx"
        headers = ["CODGEO", "LIBGEO"] + [f"M{y}" for y in range(2018, 2007, -1)]
        return self.get(remote_file_path, headers)

    def get_data(self):
        pop = self.get_pop_data()
        household = self.get_household_data()
        for insee, data in pop.items():
            data.update(household[str(insee)])
        return list(pop.values())

    def upload_pop(self):
        logger.info("Load population and household from Excel file on S3")
        logger.info("Delete previous data and reset id counter")
        TruncateComPop.truncate()
        data = self.get_data()
        logger.info("Begin looping on Excel rows")
        todo = []
        commune_list = {city.insee: city for city in Commune.objects.all()}
        for row in data:
            if not row["CODGEO"] in commune_list:
                continue

            def diff(prefix, year):
                try:
                    current = row.get(f"{prefix}{year}", None)
                    previous = row.get(f"{prefix}{year-1}", None)
                    return current - previous
                except (TypeError, ValueError):
                    return None

            todo += [
                CommunePop(
                    city=commune_list[row["CODGEO"]],
                    year=y,
                    pop=row.get(f"P{y}", None),
                    pop_change=diff("P", y),
                    household=row.get(f"M{y}", None),
                    household_change=diff("M", y),
                )
                for y in range(2019, 2005, -1)
            ]
            if len(todo) >= 10000:
                logger.info(f"Save to bdd, INSEE so far {row['CODGEO']}")
                CommunePop.objects.bulk_create(todo)
                del todo
                todo = []
